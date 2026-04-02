# -*- coding: utf-8 -*-
"""Generate literature review Excel for Jiang Hao"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ---- Literature Data ----
LITERATURE = [
    {
        "info": "[1] Vaswani A, Shazeer N, Parmar N, et al. Attention Is All You Need[C]. "
                "Advances in Neural Information Processing Systems (NeurIPS), 2017: 5998-6008.",
        "summary": "提出 Transformer 架构，以自注意力机制取代传统 RNN/CNN，"
                   "实现序列到序列建模的并行化。该架构成为后续 GPT、BERT 等大语言模型的基础，"
                   "直接决定了本项目所依赖的 LLM 推理能力的技术根基。",
        "thought": "Transformer 的自注意力机制让模型能够捕获长距离依赖，"
                   "这对公文起草中跨段落语义一致性至关重要。我在设计智能公文流水线时，"
                   "深刻体会到模型对上下文窗口的利用直接影响生成质量。"
    },
    {
        "info": "[2] Brown T B, Mann B, Ryder N, et al. Language Models are Few-Shot Learners[C]. "
                "Advances in Neural Information Processing Systems (NeurIPS), 2020: 1877-1901.",
        "summary": "展示 GPT-3 的 Few-Shot / Zero-Shot 能力，证明足够大的语言模型"
                   "可以通过少量示例完成多种下游任务，无需微调。"
                   "为本项目通过 Prompt Engineering 驱动公文起草提供了理论支撑。",
        "thought": "Few-Shot 能力让我们不必为每种公文类型训练专门模型，"
                   "而是通过精心设计的 Prompt 模板即可适配通知、报告、请示等多种文种，"
                   "极大降低了系统开发和维护成本。"
    },
    {
        "info": "[3] Lewis P, Perez E, Piktus A, et al. Retrieval-Augmented Generation for "
                "Knowledge-Intensive NLP Tasks[C]. Advances in Neural Information Processing "
                "Systems (NeurIPS), 2020: 9459-9474.",
        "summary": "提出 RAG (检索增强生成) 范式，将外部知识库检索与语言模型生成相结合，"
                   "有效减少模型幻觉并提升事实准确性。"
                   "本项目的智能问答和知识库模块直接采用了该架构思路。",
        "thought": "RAG 是解决公文场景中政策法规准确引用的关键。"
                   "实际开发中我发现，检索质量对最终回答影响巨大，"
                   "需要在分块策略、向量化模型和相似度阈值上反复调优。"
    },
    {
        "info": "[4] Wei J, Wang X, Schuurmans D, et al. Chain-of-Thought Prompting Elicits "
                "Reasoning in Large Language Models[C]. Advances in Neural Information Processing "
                "Systems (NeurIPS), 2022: 24824-24837.",
        "summary": "提出 Chain-of-Thought (CoT) 提示方法，通过在提示中加入推理链，"
                   "显著提升 LLM 在复杂推理任务上的表现。"
                   "本项目在公文审查和格式建议环节借鉴了该思路。",
        "thought": "在公文审查模块中，我借鉴 CoT 思想让模型先列出审查要点再逐项检查，"
                   "输出结构化的审查意见而非笼统结论，"
                   "用户反馈这种方式更具可解释性和可操作性。"
    },
    {
        "info": "[5] Devlin J, Chang M W, Lee K, et al. BERT: Pre-training of Deep Bidirectional "
                "Transformers for Language Understanding[C]. Proceedings of NAACL-HLT, 2019: 4171-4186.",
        "summary": "提出 BERT 预训练模型，采用双向 Transformer 编码器和掩码语言模型任务，"
                   "在多项 NLU 基准上取得突破。其向量化表示能力"
                   "为本项目知识库的文档嵌入和语义检索提供了技术参考。",
        "thought": "BERT 的双向编码思想帮助我理解了文档嵌入的原理。"
                   "在知识库检索中，高质量的文本向量化是 RAG 链路的基石，"
                   "选择合适的 Embedding 模型直接决定检索召回率。"
    },
    {
        "info": "[6] Tichy W F. RCS - A System for Version Control[J]. Software: Practice and "
                "Experience, 1985, 15(7): 637-654.",
        "summary": "讨论版本控制系统的设计原理，包括差异计算和版本追踪机制。"
                   "本项目采用 Git 进行多人协同开发，"
                   "文档版本管理思想也应用在公文的多阶段流水线设计中。",
        "thought": "多人协同开发中版本控制至关重要。项目中五人同时开发不同模块，"
                   "通过规范的 Git 分支策略和 rebase 流程，"
                   "有效避免了代码冲突和功能覆盖问题。"
    },
    {
        "info": "[7] Fielding R T. Architectural Styles and the Design of Network-based Software "
                "Architectures[D]. University of California, Irvine, 2000.",
        "summary": "定义 REST 架构风格，阐述无状态、统一接口、资源导向等约束条件。"
                   "本项目后端 API 完全遵循 RESTful 设计规范，"
                   "所有接口统一返回 {code, message, data} 结构。",
        "thought": "REST 规范让前后端职责清晰分离。"
                   "实际开发中统一的响应格式极大简化了前端错误处理逻辑，"
                   "也让 API 文档自动生成 (OpenAPI) 变得自然。"
    },
    {
        "info": "[8] Merkel D. Docker: Lightweight Linux Containers for Consistent Development "
                "and Deployment[J]. Linux Journal, 2014, 2014(239): 2.",
        "summary": "介绍 Docker 容器化技术，实现应用的环境隔离和一致性部署。"
                   "本项目采用 Docker Compose 编排 PostgreSQL、Redis、"
                   "后端、前端等 6 个容器，实现一键部署。",
        "thought": "容器化解决了团队成员开发环境不一致的痛点。"
                   "特别是字体、Chromium 等系统依赖，通过 Dockerfile 统一安装，"
                   "避免了在不同操作系统上手动配置的问题。"
    },
    {
        "info": "[9] Sandoval-Almazan R, Gil-Garcia J R. Toward an Integrative Assessment of "
                "Open Government: Proposing Conceptual Lenses and Practical Components[J]. "
                "Journal of Organizational Computing and Electronic Commerce, 2016, 26(1-2): 170-192.",
        "summary": "探讨电子政务开放治理的评估框架，分析信息技术在政府文档管理、"
                   "透明度提升和公民参与中的作用。"
                   "为本项目的政务公文智能化处理场景提供了需求层面的理论依据。",
        "thought": "这篇文献拓宽了我对政务信息化的认识。"
                   "智能公文系统不只是技术工具，还承载着提升行政效率、"
                   "降低人工审核成本的治理价值，这激励我更重视系统的规范性和可靠性。"
    },
    {
        "info": "[10] GB/T 9704-2012. 党政机关公文格式[S]. 北京: 中国标准出版社, 2012.",
        "summary": "规定党政机关公文的用纸、排版、字体字号、版头版记等格式要求。"
                   "本项目排版引擎和样式标准化算法的核心依据，"
                   "系统中仿宋三号正文、黑体二号标题等规范均源自该标准。",
        "thought": "国标 GB/T 9704 是公文排版的硬约束。"
                   "开发格式化引擎时我把每条规则逐项实现为代码规则，"
                   "确保 AI 生成内容经过标准化后的输出严格符合国家标准。"
    },
]

def build_excel(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "文献综述"

    # -- Styles --
    hdr_font = Font(name="SimHei", bold=True, size=12, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="SimSun", size=11)
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    # -- Title row --
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="GovAI 智能公文处理平台 -- 文献综述记录表 (蒋浩)")
    title_cell.font = Font(name="SimHei", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # -- Headers --
    headers = ["序号", "文献基本信息", "文献内容简介", "个人心得"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin
    ws.row_dimensions[2].height = 28

    # -- Data rows --
    even_fill = PatternFill("solid", fgColor="D6E4F0")
    for i, lit in enumerate(LITERATURE, 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i).font = cell_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin

        ws.cell(row=row, column=2, value=lit["info"]).font = cell_font
        ws.cell(row=row, column=2).alignment = cell_align
        ws.cell(row=row, column=2).border = thin

        ws.cell(row=row, column=3, value=lit["summary"]).font = cell_font
        ws.cell(row=row, column=3).alignment = cell_align
        ws.cell(row=row, column=3).border = thin

        ws.cell(row=row, column=4, value=lit["thought"]).font = cell_font
        ws.cell(row=row, column=4).alignment = cell_align
        ws.cell(row=row, column=4).border = thin

        ws.row_dimensions[row].height = 90

        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = even_fill

    # -- Column widths --
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 42

    wb.save(path)
    print(f"Saved: {path}")


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GovAI文献综述_蒋浩.xlsx")
    build_excel(out)
    print("Done!")
