#!/usr/bin/env python3
"""生成 GovAI 项目开发计划 Excel 表 + 甘特图"""

import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
import numpy as np

# ──────────── 数据定义 ────────────

MEMBERS = [
    ("蒋浩（队长）", "项目管理架构、登录注册、智能公文"),
    ("2022214419 刘硕博", "智能问答、知识图谱"),
    ("2022214408 刘瑷玮", "用户权限、知识库"),
    ("2022214414 刘聪颖", "安全规则、系统审计"),
    ("2022214423 田豪", "模型管理、用量统计"),
]

# (负责人, 任务名称, 开始日期, 结束日期, 阶段, 交付物)
TASKS = [
    # ── 第一阶段：需求分析与架构设计 ──
    ("蒋浩（队长）",       "项目架构设计与技术选型",       date(2026,3,1),  date(2026,3,3),  "需求与设计", "架构设计文档、技术路线"),
    ("蒋浩（队长）",       "数据库 Schema 设计",          date(2026,3,3),  date(2026,3,5),  "需求与设计", "schema.sql、ER 图"),
    ("2022214419 刘硕博",  "智能问答需求分析与调研",       date(2026,3,1),  date(2026,3,4),  "需求与设计", "问答模块需求文档"),
    ("2022214408 刘瑷玮",  "用户权限模型设计",            date(2026,3,1),  date(2026,3,4),  "需求与设计", "RBAC 权限矩阵"),
    ("2022214414 刘聪颖",  "安全规则需求调研",            date(2026,3,1),  date(2026,3,4),  "需求与设计", "安全规则清单"),
    ("2022214423 田豪",    "模型管理与用量需求分析",       date(2026,3,1),  date(2026,3,4),  "需求与设计", "模型管理方案"),

    # ── 第二阶段：核心模块开发 ──
    ("蒋浩（队长）",       "登录注册模块（JWT + RBAC）",   date(2026,3,5),  date(2026,3,10), "核心开发", "登录/注册 API + 前端页面"),
    ("蒋浩（队长）",       "智能公文 — 起草模块",         date(2026,3,10), date(2026,3,16), "核心开发", "起草 SSE 流式 + Dify 对接"),
    ("蒋浩（队长）",       "智能公文 — 审查 & 优化模块",   date(2026,3,16), date(2026,3,20), "核心开发", "审查/优化 API + 前端交互"),
    ("蒋浩（队长）",       "智能公文 — 排版 & 导出",      date(2026,3,20), date(2026,3,25), "核心开发", "格式化引擎 + PDF/DOCX 导出"),

    ("2022214419 刘硕博",  "智能问答前后端开发",           date(2026,3,5),  date(2026,3,14), "核心开发", "问答聊天 SSE + RAG 链路"),
    ("2022214419 刘硕博",  "知识图谱模块开发",            date(2026,3,14), date(2026,3,22), "核心开发", "Apache AGE 图谱 + 可视化"),

    ("2022214408 刘瑷玮",  "用户/角色/权限 CRUD",         date(2026,3,5),  date(2026,3,12), "核心开发", "用户管理 API + 前端页面"),
    ("2022214408 刘瑷玮",  "知识库模块开发",              date(2026,3,12), date(2026,3,22), "核心开发", "知识库上传/检索 + Dify 同步"),

    ("2022214414 刘聪颖",  "安全规则模块开发",            date(2026,3,5),  date(2026,3,14), "核心开发", "敏感词检测 + 规则配置"),
    ("2022214414 刘聪颖",  "系统审计模块开发",            date(2026,3,14), date(2026,3,22), "核心开发", "操作日志 + 审计查询"),

    ("2022214423 田豪",    "模型管理模块开发",            date(2026,3,5),  date(2026,3,14), "核心开发", "模型列表/切换/配额 API"),
    ("2022214423 田豪",    "用量统计模块开发",            date(2026,3,14), date(2026,3,22), "核心开发", "调用统计 + 图表展示"),

    # ── 第三阶段：集成测试 ──
    ("蒋浩（队长）",       "全链路集成与联调",            date(2026,3,25), date(2026,3,28), "集成测试", "端到端测试报告"),
    ("2022214419 刘硕博",  "问答 & 图谱集成测试",         date(2026,3,22), date(2026,3,27), "集成测试", "测试用例 + Bug 修复"),
    ("2022214408 刘瑷玮",  "权限 & 知识库集成测试",       date(2026,3,22), date(2026,3,27), "集成测试", "测试用例 + Bug 修复"),
    ("2022214414 刘聪颖",  "安全 & 审计集成测试",         date(2026,3,22), date(2026,3,27), "集成测试", "安全验收报告"),
    ("2022214423 田豪",    "模型 & 统计集成测试",         date(2026,3,22), date(2026,3,27), "集成测试", "测试用例 + Bug 修复"),

    # ── 第四阶段：部署与交付 ──
    ("蒋浩（队长）",       "Docker 部署与文档编写",        date(2026,3,28), date(2026,3,31), "部署交付", "部署指南 + 项目文档"),
    ("2022214419 刘硕博",  "问答模块文档与优化",           date(2026,3,27), date(2026,3,31), "部署交付", "模块说明文档"),
    ("2022214408 刘瑷玮",  "权限模块文档与优化",           date(2026,3,27), date(2026,3,31), "部署交付", "模块说明文档"),
    ("2022214414 刘聪颖",  "安全模块文档与优化",           date(2026,3,27), date(2026,3,31), "部署交付", "安全说明文档"),
    ("2022214423 田豪",    "统计模块文档与优化",           date(2026,3,27), date(2026,3,31), "部署交付", "模块说明文档"),
]

# ──────────── Excel 生成 ────────────

def build_excel(path: str):
    wb = Workbook()

    # ── Sheet 1: 团队成员 ──
    ws_team = wb.active
    ws_team.title = "团队成员"
    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=11)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    team_headers = ["序号", "姓名（学号）", "角色", "负责模块"]
    for col, h in enumerate(team_headers, 1):
        c = ws_team.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border
    for i, (name, modules) in enumerate(MEMBERS, 1):
        role = "队长" if i == 1 else "开发"
        row = [i, name, role, modules]
        for col, val in enumerate(row, 1):
            c = ws_team.cell(row=i+1, column=col, value=val)
            c.font = cell_font; c.alignment = cell_align; c.border = thin_border
    ws_team.column_dimensions["A"].width = 6
    ws_team.column_dimensions["B"].width = 24
    ws_team.column_dimensions["C"].width = 8
    ws_team.column_dimensions["D"].width = 36

    # ── Sheet 2: 开发计划 ──
    ws_plan = wb.create_sheet("开发计划")
    plan_headers = ["序号", "阶段", "负责人", "任务名称", "开始日期", "结束日期", "工期(天)", "交付物"]
    for col, h in enumerate(plan_headers, 1):
        c = ws_plan.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

    phase_fills = {
        "需求与设计": PatternFill("solid", fgColor="D6E4F0"),
        "核心开发":   PatternFill("solid", fgColor="E2EFDA"),
        "集成测试":   PatternFill("solid", fgColor="FFF2CC"),
        "部署交付":   PatternFill("solid", fgColor="FCE4D6"),
    }

    for i, (person, task, start, end, phase, deliverable) in enumerate(TASKS, 1):
        duration = (end - start).days
        row_data = [i, phase, person, task, start, end, duration, deliverable]
        fill = phase_fills.get(phase)
        for col, val in enumerate(row_data, 1):
            c = ws_plan.cell(row=i+1, column=col, value=val)
            c.font = cell_font; c.alignment = cell_align; c.border = thin_border
            if fill:
                c.fill = fill
            if isinstance(val, date):
                c.number_format = "YYYY-MM-DD"

    ws_plan.column_dimensions["A"].width = 6
    ws_plan.column_dimensions["B"].width = 12
    ws_plan.column_dimensions["C"].width = 22
    ws_plan.column_dimensions["D"].width = 34
    ws_plan.column_dimensions["E"].width = 14
    ws_plan.column_dimensions["F"].width = 14
    ws_plan.column_dimensions["G"].width = 10
    ws_plan.column_dimensions["H"].width = 32

    # ── Sheet 3: 甘特图数据（日历视图） ──
    ws_gantt = wb.create_sheet("甘特图(日历)")
    # 行: 任务名称 | 3/1 | 3/2 | ... | 3/31
    ws_gantt.cell(row=1, column=1, value="任务").font = header_font
    ws_gantt.cell(row=1, column=1).fill = header_fill
    ws_gantt.cell(row=1, column=1).alignment = header_align
    ws_gantt.cell(row=1, column=1).border = thin_border

    ws_gantt.cell(row=1, column=2, value="负责人").font = header_font
    ws_gantt.cell(row=1, column=2).fill = header_fill
    ws_gantt.cell(row=1, column=2).alignment = header_align
    ws_gantt.cell(row=1, column=2).border = thin_border

    for d in range(31):
        day = date(2026, 3, 1) + timedelta(days=d)
        c = ws_gantt.cell(row=1, column=d+3, value=f"{day.month}/{day.day}")
        c.font = Font(name="微软雅黑", bold=True, size=9, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    bar_fills = {
        "需求与设计": PatternFill("solid", fgColor="4472C4"),
        "核心开发":   PatternFill("solid", fgColor="548235"),
        "集成测试":   PatternFill("solid", fgColor="BF8F00"),
        "部署交付":   PatternFill("solid", fgColor="C55A11"),
    }

    for i, (person, task, start, end, phase, _) in enumerate(TASKS, 1):
        row = i + 1
        ws_gantt.cell(row=row, column=1, value=task).font = Font(name="微软雅黑", size=9)
        ws_gantt.cell(row=row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
        ws_gantt.cell(row=row, column=1).border = thin_border

        ws_gantt.cell(row=row, column=2, value=person.split("）")[-1] if "）" in person else person.split(" ")[-1]).font = Font(name="微软雅黑", size=9)
        ws_gantt.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=row, column=2).border = thin_border

        for d in range(31):
            day = date(2026, 3, 1) + timedelta(days=d)
            c = ws_gantt.cell(row=row, column=d+3)
            c.border = thin_border
            if start <= day < end:
                c.fill = bar_fills.get(phase, PatternFill("solid", fgColor="4472C4"))

    ws_gantt.column_dimensions["A"].width = 30
    ws_gantt.column_dimensions["B"].width = 10
    for d in range(31):
        ws_gantt.column_dimensions[get_column_letter(d+3)].width = 4.5

    wb.save(path)
    print(f"Excel saved: {path}")


# ──────────── 甘特图 PNG ────────────

def build_gantt_png(path: str):
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial"]
    plt.rcParams["axes.unicode_minus"] = False

    phase_colors = {
        "需求与设计": "#4472C4",
        "核心开发":   "#548235",
        "集成测试":   "#BF8F00",
        "部署交付":   "#C55A11",
    }

    fig, ax = plt.subplots(figsize=(20, 14))

    # 反转任务顺序，让第一个任务在最上面
    tasks_reversed = list(reversed(TASKS))
    n = len(tasks_reversed)

    for i, (person, task, start, end, phase, _) in enumerate(tasks_reversed):
        duration = (end - start).days
        short_name = person.split("）")[-1] if "）" in person else person.split(" ")[-1]
        label = f"{short_name} | {task}"
        color = phase_colors.get(phase, "#4472C4")

        ax.barh(i, duration, left=mdates.date2num(start), height=0.6,
                color=color, edgecolor="white", linewidth=0.5, alpha=0.9)

        # 在条形中间写天数
        mid = mdates.date2num(start) + duration / 2
        ax.text(mid, i, f"{duration}d", ha="center", va="center",
                fontsize=8, color="white", fontweight="bold")

    ax.set_yticks(range(n))
    ax.set_yticklabels([
        f"{t[1]}" for t in tasks_reversed
    ], fontsize=9)

    ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
    ax.set_xlim(mdates.date2num(date(2026,2,28)), mdates.date2num(date(2026,4,1)))
    plt.xticks(rotation=45, fontsize=8)

    # 阶段分隔线
    # 找到各阶段的边界
    phase_boundaries = []
    prev_phase = None
    for i, t in enumerate(tasks_reversed):
        if t[4] != prev_phase:
            if prev_phase is not None:
                phase_boundaries.append(i - 0.5)
            prev_phase = t[4]

    for b in phase_boundaries:
        ax.axhline(y=b, color="#cccccc", linewidth=0.8, linestyle="--")

    # 图例
    from matplotlib.patches import Patch
    legend_items = [Patch(facecolor=c, label=p) for p, c in phase_colors.items()]
    ax.legend(handles=legend_items, loc="upper right", fontsize=10, framealpha=0.9)

    # 周末底色
    for d in range(31):
        day = date(2026, 3, 1) + timedelta(days=d)
        if day.weekday() >= 5:  # 周六日
            ax.axvspan(mdates.date2num(day), mdates.date2num(day + timedelta(days=1)),
                       alpha=0.08, color="gray", zorder=0)

    ax.set_title("GovAI 智能公文处理平台 — 项目开发甘特图（2026年3月）",
                 fontsize=16, fontweight="bold", pad=15)
    ax.set_xlabel("日期", fontsize=11)
    ax.grid(axis="x", alpha=0.3, linewidth=0.5)
    ax.invert_yaxis()

    # 右侧标注负责人
    ax2 = ax.twinx()
    ax2.set_ylim(ax.get_ylim())
    ax2.set_yticks(range(n))
    ax2.set_yticklabels([
        (t[0].split("）")[-1] if "）" in t[0] else t[0].split(" ")[-1])
        for t in tasks_reversed
    ], fontsize=9, color="#666666")

    plt.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Gantt chart saved: {path}")


if __name__ == "__main__":
    out_dir = os.path.dirname(os.path.abspath(__file__))
    build_excel(os.path.join(out_dir, "GovAI项目开发计划.xlsx"))
    build_gantt_png(os.path.join(out_dir, "GovAI项目甘特图.png"))
    print("Done!")
